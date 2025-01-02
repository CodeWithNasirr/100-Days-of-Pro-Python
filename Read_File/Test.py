
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QMessageBox
import os
from openpyxl import load_workbook
import openpyxl
import openpyxl.workbook
import platform
from datetime import datetime

class Ui_MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi()

    def setupUi(self):
        self.setObjectName("MainWindow")
        self.setEnabled(True)
        self.resize(390, 535)
        self.setMaximumSize(QtCore.QSize(390, 535))
        self.setStyleSheet("""
                QWidget#centralwidget {
                    background-color: qconicalgradient(cx:0.5, cy:0.5, angle:0, stop:0 rgba(35, 40, 3, 255), stop:0.16 rgba(136, 106, 22, 255), stop:0.225 rgba(166, 140, 41, 255), stop:0.285 rgba(204, 181, 74, 255), stop:0.345 rgba(235, 219, 102, 255), stop:0.415 rgba(245, 236, 112, 255), stop:0.52 rgba(209, 190, 76, 255), stop:0.57 rgba(187, 156, 51, 255), stop:0.635 rgba(168, 142, 42, 255), stop:0.695 rgba(202, 174, 68, 255), stop:0.75 rgba(218, 202, 86, 255), stop:0.815 rgba(208, 187, 73, 255), stop:0.88 rgba(187, 156, 51, 255), stop:0.935 rgba(137, 108, 26, 255), stop:1 rgba(35, 40, 3, 255));}""")

        # Central Widget Open
        self.centralwidget = QtWidgets.QWidget(self)
        self.centralwidget.setObjectName("centralwidget")
        #Heading
        self.Heading = QtWidgets.QLabel(self.centralwidget)
        self.Heading.setGeometry(QtCore.QRect(50, 30, 309, 24))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Heading.setFont(font)
        self.Heading.setObjectName("Heading")
        # self.Heading.setStyleSheet("color: rgb(209, 190, 76);")  # Example solid color

        #Heading2
        self.Heading_2 = QtWidgets.QLabel(self.centralwidget)
        self.Heading_2.setGeometry(QtCore.QRect(80, 60, 241, 41))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.Heading_2.setFont(font)
        self.Heading_2.setObjectName("Heading_2")

        # First_Name
        self.F_Name = QtWidgets.QLabel(self.centralwidget)
        self.F_Name.setGeometry(QtCore.QRect(40, 120, 82, 31))
        self.F_Name.setObjectName("F_Name")

        # First_Name line Edit
        self.F_Name_lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.F_Name_lineEdit.setGeometry(QtCore.QRect(130, 120, 211, 30))
        self.F_Name_lineEdit.setObjectName("F_Name_lineEdit")
        self.F_Name_lineEdit.setStyleSheet('border-radius:15px; font:25 10pt Calibri;')
        self.F_Name_lineEdit.setPlaceholderText('Enter Your First Name')

        # Last_Name
        self.L_Name = QtWidgets.QLabel(self.centralwidget)
        self.L_Name.setGeometry(QtCore.QRect(40, 170, 82, 31))
        self.L_Name.setObjectName("F_Name")

        #Last_Name Edit
        self.L_Name_input = QtWidgets.QLineEdit(self.centralwidget)
        self.L_Name_input.setGeometry(QtCore.QRect(130, 170, 211, 30))
        self.L_Name_input.setObjectName("L_Name_input")
        self.L_Name_input.setStyleSheet('border-radius:15px; font:25 10pt Calibri;')
        self.L_Name_input.setPlaceholderText('Enter Your Last Name')

        # Email
        self.Email = QtWidgets.QLabel(self.centralwidget)
        self.Email.setGeometry(QtCore.QRect(40, 220, 82, 31))
        self.Email.setObjectName("Email")

        # Email Edit
        self.Email_input = QtWidgets.QLineEdit(self.centralwidget)
        self.Email_input.setGeometry(QtCore.QRect(130, 220, 211, 30))
        self.Email_input.setObjectName("Email_input")
        self.Email_input.setStyleSheet('border-radius:15px; font:25 10pt Calibri;')
        self.Email_input.setPlaceholderText('Enter Your Email')


        # Number
        self.Number = QtWidgets.QLabel(self.centralwidget)
        self.Number.setGeometry(QtCore.QRect(40, 270, 82, 31))
        self.Number.setObjectName("Number")

        # Number input
        self.Number_input = QtWidgets.QLineEdit(self.centralwidget)
        self.Number_input.setGeometry(QtCore.QRect(130, 270, 211, 30))
        self.Number_input.setObjectName("Number_input")
        self.Number_input.setStyleSheet('border-radius:15px; font:25 10pt Calibri;')
        self.Number_input.setPlaceholderText('Enter Your Number')


        # Company
        self.Company = QtWidgets.QLabel(self.centralwidget)
        self.Company.setGeometry(QtCore.QRect(40, 320, 82, 31))
        self.Company.setObjectName("Company")

        # Company input
        self.Company_input = QtWidgets.QLineEdit(self.centralwidget)
        self.Company_input.setGeometry(QtCore.QRect(130, 320, 211, 30))
        self.Company_input.setObjectName("Company_input")
        self.Company_input.setStyleSheet('border-radius:15px; font:25 10pt Calibri;')
        self.Company_input.setPlaceholderText('Enter Company Name')

        

        # Category
        self.Category = QtWidgets.QLabel(self.centralwidget)
        self.Category.setGeometry(QtCore.QRect(40, 360, 82, 41))
        self.Category.setObjectName("Category")

        # Vip input
        self.Vip = QtWidgets.QRadioButton(self.centralwidget)
        self.Vip.setGeometry(QtCore.QRect(100, 360, 91, 51))
        self.Vip.setObjectName("Vip")
        self.Vip.setStyleSheet("font: 20pt \"Arial Rounded MT Bold\";\n"
"color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(0, 0, 0, 255), stop:0.05 rgba(14, 8, 73, 255), stop:0.36 rgba(28, 17, 145, 255), stop:0.6 rgba(126, 14, 81, 255), stop:0.75 rgba(234, 11, 11, 255), stop:0.79 rgba(244, 70, 5, 255), stop:0.86 rgba(255, 136, 0, 255), stop:0.935 rgba(239, 236, 55, 255));")
        
        # Normal input
        self.Normal = QtWidgets.QRadioButton(self.centralwidget)
        self.Normal.setGeometry(QtCore.QRect(180, 360, 131, 51))
        self.Normal.setObjectName("Normal")
        self.Normal.setStyleSheet("font: 20pt \"Arial Rounded MT Bold\";\n"
"color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(0, 0, 0, 255), stop:0.05 rgba(14, 8, 73, 255), stop:0.36 rgba(28, 17, 145, 255), stop:0.6 rgba(126, 14, 81, 255), stop:0.75 rgba(234, 11, 11, 255), stop:0.79 rgba(244, 70, 5, 255), stop:0.86 rgba(255, 136, 0, 255), stop:0.935 rgba(239, 236, 55, 255));")
        


        #Submit button
        self.Sumbit_Button = QtWidgets.QPushButton(self.centralwidget)
        self.Sumbit_Button.setGeometry(QtCore.QRect(90, 440, 75, 41))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(10)
        self.Sumbit_Button.setFont(font)
        self.Sumbit_Button.setStyleSheet("border-radius:20px;background-color: rgb(69, 138, 0);color:rgb(255, 255, 255)")
        self.Sumbit_Button.setObjectName("Sumbit_Button")
        self.Sumbit_Button.clicked.connect(self.Submit)
        

        #Clear() button
        self.Clear_Button = QtWidgets.QPushButton(self.centralwidget)
        self.Clear_Button.setGeometry(QtCore.QRect(210, 440, 75, 41))
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(10)
        self.Clear_Button.setFont(font)
        self.Clear_Button.setStyleSheet("border-radius:20px;background-color: rgb(193, 70, 46);color:rgb(255, 255, 255)")
        self.Clear_Button.setObjectName("Clear_Button")
        self.Clear_Button.clicked.connect(self.clear)



        # Central Widget End
        self.setCentralWidget(self.centralwidget)

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("MainWindow","MainWindow"))
        self.Heading.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color: rgb(75, 46, 46);\">Indian Avenger Seminar-2025</span></p></body></html>"))
        self.Heading_2.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-weight:600; color: rgb(75, 46, 46);\">Indore Jan 9th to 10th</span></p></body></html>"))
        self.F_Name.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">First Name :</span></p></body></html>"))
        self.L_Name.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Last Name :</span></p></body></html>"))
        self.Email.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Email :</span></p></body></html>"))
        self.Number.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Number :</span></p></body></html>"))
        self.Company.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Company :</span></p></body></html>"))
        # self.Category.setText(_translate("MainWindow", "<html><head/><body><p align=\"center\"><span style=\" font-size:10pt; font-weight:600; color:#ffffff;\">Category :</span></p></body></html>"))
        self.Vip.setText(_translate("MainWindow", "VIP"))
        self.Normal.setText(_translate("MainWindow", "Normal"))
        self.Sumbit_Button.setText(_translate("MainWindow", "Submit"))
        self.Clear_Button.setText(_translate("MainWindow", "Clear"))



    def show_message(self, title,text,icon):
        msg = QMessageBox()
        msg.setIcon(icon)
        msg.setWindowTitle(title)
        msg.setText(text)
        msg.setStyleSheet("background:white;color:black")
        msg.exec_()

    def clear(self):
        self.F_Name_lineEdit.clear()
        self.L_Name_input.clear()
        self.Number_input.clear()
        self.Email_input.clear()
        self.Company_input.clear()

        
    
    def email_valid(self,email):#skofficial665@gmail.com
        k,j,f=0,0,0
        if len(email)>=6:
            if email[0].isalpha():
                if ("@" in email) and email.count("@")==1:
                    if email[-4]=="." or email[-3]==".":
                        if "gmail" in email:
                            for i in email:
                                if i==i.isspace():
                                    k=1
                                elif i.isalpha():
                                    if i.isupper():
                                        j=1
                                elif i.isdigit():
                                    continue
                                elif i=="_" or i=="." or i=="@":
                                    continue
                                else:
                                    f=1

                            if k==1 or j==1 or f==1:
                                return False
                            else:
                                return True
                        else:
                            return False
                    else:
                        return False
                else:
                    return False
            else:
                return False
        else:
            return False
    
    def valid_Number(self,input_value):
        try:
            float(input_value)
            return True
        except ValueError:
            return False
    
    def open_explorer(self,path):
        system_plateform=platform.system()
        if system_plateform =="Windows":
            os.startfile(path)
        elif system_plateform == "Linux":
            os.startfile(path)
        #ADD MORE OS
        else:
            self.show_message("Invalid OS")
    
    def Submit(self):
        first_name=self.F_Name_lineEdit.text()
        last_name=self.L_Name_input.text()
        PhoneNumber=self.Number_input.text()
        Email=self.Email_input.text()
        Company=self.Company_input.text()
        Category=None

        if not first_name:
            self.show_message('Error',"Enter FirstName",QMessageBox.Warning)
            return

        elif not last_name:
            self.show_message('Error',"Enter LastName",QMessageBox.Warning)
            return

        if not Email:
            self.show_message("Error",f"Enter a email id ",QMessageBox.Warning)
            return
        elif not self.email_valid(Email):
            self.show_message("Error",f"{Email} is not a valid email.",QMessageBox.Warning)
            return
        if not PhoneNumber:
            self.show_message('Error',f"Enter Phone Number",QMessageBox.Warning)
            return
        elif self.valid_Number(PhoneNumber) is False:
            self.show_message('Error',f"Enter a valid Number",QMessageBox.Warning)
            return
        if not Company:
            self.show_message('Error',f"Enter Company Name",QMessageBox.Warning)
            return

        if not (self.Vip.isChecked() or self.Normal.isChecked()):
            self.show_message('Error',f"Select Category ",QMessageBox.Warning)
            return

        if self.Normal.isChecked():
            Category="Normal"
   
        elif self.Vip.isChecked():
            Category="Vip"
      
        location=r'C:\Seminardata'
        if not os.path.exists(location):
            os.makedirs(location)

        file_path='Seminar_Data.xlsx'
        target_dic=os.path.join(location,file_path)
        if not os.path.exists(target_dic):
            wb=openpyxl.Workbook()
            sheet=wb.active
            Heading=["First_name",
                    'Last_name',
                    'PhoneNumber',
                    'Email',
                    'Company',
                    'Category',
                    "DATE-TIME"]
            sheet.append(Heading)
            wb.save(target_dic)
        if target_dic:
            wb=openpyxl.load_workbook(target_dic)
            sheet=wb.active
            current = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            sheet.append([first_name,last_name,PhoneNumber,Email,Company,Category,current])
            wb.save(target_dic)
            self.show_message("Saved",f"Data Saved to {target_dic}",QMessageBox.Information)

        if os.path.exists(target_dic):
            self.open_explorer(os.path.dirname(target_dic))
        else:
            self.show_message("Error", "File copy Failed")

    
if __name__ == "__main__":
    import sys
    app = QApplication(sys.argv)
    ui = Ui_MainWindow()
    ui.show()
    sys.exit(app.exec_())
